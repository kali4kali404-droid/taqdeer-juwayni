from fastapi import FastAPI, APIRouter, HTTPException, Depends, Query, Response, Body, Request
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.middleware.cors import CORSMiddleware
from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient
import os
import uuid
import bcrypt
import jwt
from jwt.exceptions import ExpiredSignatureError, InvalidTokenError
from datetime import datetime, timezone, timedelta
from pathlib import Path
from pydantic import BaseModel
from typing import List, Optional
import io

# ================== INIT ==================
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

mongo_url = os.getenv("MONGO_URL", "mongodb://localhost:27017")
client = AsyncIOMotorClient(mongo_url)
db = client[os.getenv("DB_NAME", "taqdir_db")]

SECRET_KEY = os.getenv("JWT_SECRET")
if not SECRET_KEY:
    raise RuntimeError("❌ JWT_SECRET غير موجود في ملف .env")

ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_HOURS = 24
ALLOWED_ORIGINS = os.getenv("ALLOWED_ORIGINS", "http://localhost:3000").split(",")

app = FastAPI(title="منصة ثانوية الإمام الجويني الإلكترونية لاختبارات - القدرات - التحصيلي")

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

api_router = APIRouter(prefix="/api")
security = HTTPBearer()


# ================== MODELS ==================
class UserLogin(BaseModel):
    username: str
    password: str

class UserCreate(BaseModel):
    username: str
    password: str
    full_name: str
    role: str

class UserUpdate(BaseModel):
    username: str
    full_name: str
    role: str

class StudentLogin(BaseModel):
    code: str
    student_name: str
    grade: str
    class_name: Optional[str] = None
    school_name: Optional[str] = ""  # إضافة اسم المدرسة لنموذج التحقق ليكون اختيارياً

class StudentCodeRequest(BaseModel):
    code: Optional[str] = None

class StudentExamStart(BaseModel):
    exam_id: str
    student_name: Optional[str] = None  
    student_phone: Optional[str] = None
    grade: Optional[str] = None
    class_name: Optional[str] = None

class StudentAnswerSubmit(BaseModel):
    student_exam_id: str
    question_id: str
    selected_option_id: str

class AssignGradeRequest(BaseModel):
    student_exam_id: str
    score: float
    release_grade: Optional[bool] = False

class SectionModel(BaseModel):
    section_number: int
    title: str
    duration_minutes: int

class ExamCreate(BaseModel):
    title: str
    description: Optional[str] = ""
    instructions: Optional[str] = ""
    image: Optional[str] = None          # base64
    sections: Optional[List[SectionModel]] = []

class SectionTimeUpdate(BaseModel):
    exam_id: str
    section_number: int
    duration_minutes: int

class OptionModel(BaseModel):
    id: str
    text: str
    is_correct: bool

class QuestionCreate(BaseModel):
    text: str
    exam_id: str
    section_number: int
    points: Optional[int] = 1
    image: Optional[str] = None          # base64
    hint: Optional[str] = None           # 💡 تلميح للطالب
    options: List[OptionModel]

class QuestionUpdate(BaseModel):
    text: str
    section_number: int
    points: Optional[int] = 1
    image: Optional[str] = None
    hint: Optional[str] = None
    options: List[OptionModel]

class AnswerModel(BaseModel):
    question_id: str
    selected_option_id: str

class SectionSubmit(BaseModel):
    session_id: str
    section_number: int
    answers: List[AnswerModel]

class DeleteResultsRequest(BaseModel):
    exam_id: Optional[str] = None        # لو فاضي يحذف الكل


# ================== AUTH HELPERS ==================
def hash_password(p: str) -> str:
    return bcrypt.hashpw(p.encode(), bcrypt.gensalt()).decode()

def verify_password(p: str, h: str) -> bool:
    return bcrypt.checkpw(p.encode(), h.encode())

def create_token(data: dict) -> str:
    payload = data.copy()
    payload["exp"] = datetime.now(timezone.utc) + timedelta(hours=ACCESS_TOKEN_EXPIRE_HOURS)
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


# ================== USER MIDDLEWARE ==================
async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    token = credentials.credentials
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
    except ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="انتهت صلاحية الجلسة")
    except InvalidTokenError:
        raise HTTPException(status_code=401, detail="رمز المصادقة غير صالح")

    user_id = payload.get("sub")
    role = payload.get("role")

    if not user_id or not role:
        raise HTTPException(status_code=401, detail="بيانات الرمز غير مكتملة")

    if role == "student":
        return {"id": user_id, "role": "student"}

    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    if not user:
        raise HTTPException(status_code=401, detail="المستخدم غير موجود")

    return user

def require_staff(user: dict):
    if user["role"] not in ("admin", "teacher"):
        raise HTTPException(status_code=403, detail="غير مصرح")


# ================== AUTH ==================
@api_router.post("/auth/login")
async def login(data: UserLogin):
    user = await db.users.find_one({"username": data.username})
    if not user or not verify_password(data.password, user["password"]):
        raise HTTPException(status_code=401, detail="اسم المستخدم أو كلمة المرور غير صحيحة")

    token = create_token({"sub": user["id"], "role": user["role"]})
    return {
        "access_token": token,
        "user": {
            "id": user["id"],
            "username": user["username"],
            "role": user["role"],
            "full_name": user.get("full_name", "")
        }
    }

@api_router.get("/auth/me")
async def me(user=Depends(get_current_user)):
    return user

@api_router.post("/auth/student-login")
async def student_login(data: StudentLogin):
    code_text = data.code.strip()
    student_name = data.student_name.strip()
    grade = data.grade.strip()
    school_name = data.school_name.strip() if data.school_name else ""  # قراءة اسم المدرسة بدقة
    
    class_name = data.class_name.strip() if (data.class_name and data.class_name.strip()) else "عام"

    if not code_text or not student_name or not grade:
        raise HTTPException(status_code=400, detail="الرجاء إكمال جميع البيانات المطلوبة")

    code_doc = await db.student_codes.find_one({"code": code_text, "active": True})
    if not code_doc:
        raise HTTPException(status_code=401, detail="الكود غير صالح أو غير مفعّل")

    session_query = {
        "code": code_text,
        "student_name": student_name,
        "grade": grade,
        "class_name": class_name,
        "status": "in_progress"
    }

    student_id = str(uuid.uuid4())
    session_id = str(uuid.uuid4())
    
    await db.student_sessions.update_one(
        session_query,
        {"$setOnInsert": {
            "id": session_id,
            "student_id": student_id,
            "school_name": school_name,  # حفظ اسم المدرسة المدخل بالملي
            "completed_sections": [],
            "answers": [],
            "started_at": datetime.now(timezone.utc).isoformat(),
            "current_section": 1
        }},
        upsert=True
    )

    session = await db.student_sessions.find_one(session_query)
    
    token = create_token({"sub": session["student_id"], "role": "student"})
    return {
        "access_token": token,
        "user": {
            "id": session["student_id"],
            "role": "student",
            "session_id": session["id"],
            "student_name": session["student_name"],
            "grade": session["grade"],
            "class_name": session["class_name"],
            "school_name": session.get("school_name", "")  # إرجاع مدرسة الطالب بدقة
        }
    }


# ================== USERS ==================
@api_router.get("/users")
async def users(user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="للمدير فقط")
    return await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)

@api_router.post("/users")
async def create_user(data: UserCreate, user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="للمدير فقط")
    
    existing = await db.users.find_one({"username": data.username})
    if existing:
        raise HTTPException(status_code=400, detail="اسم المستخدم موجود مسبقاً")

    new_user = {
        "id": str(uuid.uuid4()),
        "username": data.username,
        "full_name": data.full_name,
        "role": data.role,
        "password": hash_password(data.password),
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(new_user)
    new_user.pop("_id", None)
    new_user.pop("password", None)
    return new_user

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, data: UserUpdate, user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="للمدير فقط")
    
    target_user = await db.users.find_one({"id": user_id})
    if not target_user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    # 🔒 حماية وقائية تمنع تعديل اسم حساب n.vf11 أو دوره الإشرافي
    if target_user.get("username") == "n.vf11" and (data.username != "n.vf11" or data.role != "admin"):
        raise HTTPException(status_code=403, detail="لا يمكن تعديل اسم المستخدم أو دور المشرف الرئيسي")

    existing = await db.users.find_one({"username": data.username, "id": {"$ne": user_id}})
    if existing:
        raise HTTPException(status_code=400, detail="اسم المستخدم مستخدم مسبقاً")

    result = await db.users.update_one({"id": user_id}, {"$set": {
        "username": data.username,
        "full_name": data.full_name,
        "role": data.role,
    }})
    return {"message": "تم تحديث المستخدم بنجاح"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="للمدير فقط")
    
    target_user = await db.users.find_one({"id": user_id})
    if not target_user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    # 🔒 حماية صارمة تمنع حذف المشرف الرئيسي n.vf11 نهائياً من قاعدة البيانات
    if target_user.get("username") == "n.vf11":
        raise HTTPException(status_code=403, detail="لا يمكن حذف حساب المشرف الرئيسي الدائم")

    result = await db.users.delete_one({"id": user_id})
    return {"message": "تم حذف المستخدم بنجاح"}


# ================== STUDENT CODES ==================
@api_router.get("/admin/student-codes")
async def get_student_codes(user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="للمدير فقط")
    codes = await db.student_codes.find({}, {"_id": 0}).to_list(1000)
    return codes

@api_router.post("/admin/student-code")
async def create_student_code(data: Optional[StudentCodeRequest] = None, user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="للمدير فقط")
    
    code_text = data.code.strip() if (data and data.code) else ""
    if not code_text:
        import random, string
        code_text = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
    
    existing = await db.student_codes.find_one({"code": code_text})
    if existing:
        raise HTTPException(status_code=400, detail="كود الدخول هذا مسجل مسبقاً")

    new_code = {
        "id": str(uuid.uuid4()),
        "code": code_text,
        "active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.student_codes.insert_one(new_code)
    new_code.pop("_id", None)
    return {"message": "تم إنشاء كود الدخول بنجاح", "code": new_code}

@api_router.put("/admin/student-codes/{code_id}")
async def update_student_code(code_id: str, data: StudentCodeRequest, user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="للمدير فقط")
    
    new_code_text = data.code.strip() if data.code else ""
    if not new_code_text:
        raise HTTPException(status_code=400, detail="لا يمكن أن يكون الكود فارغاً")

    existing = await db.student_codes.find_one({"code": new_code_text, "id": {"$ne": code_id}})
    if existing:
        raise HTTPException(status_code=400, detail="كود الدخول مستخدم مسبقاً")

    result = await db.student_codes.update_one(
        {"id": code_id},
        {"$set": {"code": new_code_text}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="كود الدخول غير موجود")
    return {"message": "تم تعديل الكود بنجاح"}

@api_router.put("/admin/student-codes/{code_id}/toggle")
async def toggle_student_code(code_id: str, user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="للمدير فقط")
    
    code_doc = await db.student_codes.find_one({"id": code_id})
    if not code_doc:
        raise HTTPException(status_code=404, detail="كود الدخول غير موجود")
    
    new_status = not code_doc.get("active", True)
    await db.student_codes.update_one(
        {"id": code_id},
        {"$set": {"active": new_status}}
    )
    status_text = "تنشيط" if new_status else "تعطيل"
    return {"message": f"تم {status_text} الكود بنجاح", "active": new_status}

@api_router.delete("/admin/student-codes/{code_id}")
async def delete_student_code(code_id: str, user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="للمدير فقط")
    result = await db.student_codes.delete_one({"id": code_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="كود الدخول غير موجود")
    return {"message": "تم حذف كود الدخول بنجاح"}


# ================== DASHBOARD STATS (حساب الطلاب والنشاط الفعلي فقط) ==================
@api_router.get("/dashboard/stats")
async def dashboard_stats(user=Depends(get_current_user)):
    total_exams       = await db.exams.count_documents({})
    published_exams   = await db.exams.count_documents({"status": "published"})
    total_questions   = await db.questions.count_documents({})
    
    total_students    = await db.student_sessions.count_documents({"exam_id": {"$exists": True, "$ne": None}})
    total_submissions = await db.student_sessions.count_documents({"status": "completed", "exam_id": {"$exists": True, "$ne": None}})
    in_progress       = await db.student_sessions.count_documents({"status": "in_progress", "exam_id": {"$exists": True, "$ne": None}})

    return {
        "total_exams": total_exams,
        "published_exams": published_exams,
        "total_questions": total_questions,
        "total_students": total_students,
        "total_submissions": total_submissions,
        "completed_submissions": total_submissions,
        "in_progress": in_progress,
    }


# ================== EXAMS ==================
@api_router.get("/exams")
async def get_exams(user=Depends(get_current_user)):
    exams = await db.exams.find({}, {"_id": 0}).to_list(1000)
    return exams

@api_router.get("/exams/{exam_id}")
async def get_exam(exam_id: str, user=Depends(get_current_user)):
    exam = await db.exams.find_one({"id": exam_id}, {"_id": 0})
    if not exam:
        raise HTTPException(status_code=404, detail="الاختبار غير موجود")
    return exam

@api_router.post("/exams")
async def create_exam(data: ExamCreate, user=Depends(get_current_user)):
    require_staff(user)
    exam = {
        "id": str(uuid.uuid4()),
        "title": data.title,
        "description": data.description,
        "instructions": data.instructions,
        "image": data.image,
        "sections": [s.model_dump() for s in data.sections],
        "status": "draft",
        "created_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.exams.insert_one(exam)
    exam.pop("_id", None)
    return exam

@api_router.put("/exams/{exam_id}")
async def update_exam(exam_id: str, data: ExamCreate, user=Depends(get_current_user)):
    require_staff(user)
    exam = await db.exams.find_one({"id": exam_id})
    if not exam:
        raise HTTPException(status_code=404, detail="الاختبار غير موجود")

    await db.exams.update_one({"id": exam_id}, {"$set": {
        "title": data.title,
        "description": data.description,
        "instructions": data.instructions,
        "image": data.image,
        "sections": [s.model_dump() for s in data.sections],
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }})
    return {"message": "تم تحديث الاختبار"}

@api_router.delete("/exams/{exam_id}")
async def delete_exam(exam_id: str, user=Depends(get_current_user)):
    require_staff(user)
    result = await db.exams.delete_one({"id": exam_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="الاختبار غير موجود")
    return {"message": "تم حذف الاختبار"}

# ✅ الاختبار يقدر يفتح ويقفل أكثر من مرة
@api_router.post("/exams/{exam_id}/publish")
async def publish_exam(exam_id: str, user=Depends(get_current_user)):
    require_staff(user)
    result = await db.exams.update_one({"id": exam_id}, {"$set": {"status": "published"}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="الاختبار غير موجود")
    return {"message": "تم نشر الاختبار"}

@api_router.post("/exams/{exam_id}/close")
async def close_exam(exam_id: str, user=Depends(get_current_user)):
    require_staff(user)
    result = await db.exams.update_one({"id": exam_id}, {"$set": {"status": "closed"}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="الاختبار غير موجود")
    return {"message": "تم إغلاق الاختبار"}

@api_router.put("/exams/section-time")
async def update_section_time(data: SectionTimeUpdate, user=Depends(get_current_user)):
    require_staff(user)
    exam = await db.exams.find_one({"id": data.exam_id})
    if not exam:
        raise HTTPException(status_code=404, detail="الاختبار غير موجود")

    sections = exam.get("sections", [])
    for s in sections:
        if s["section_number"] == data.section_number:
            s["duration_minutes"] = data.duration_minutes

    await db.exams.update_one({"id": data.exam_id}, {"$set": {"sections": sections}})
    return {"message": "تم تحديث وقت القسم"}


# ================== QUESTIONS ==================
@api_router.get("/questions")
async def get_questions(
    exam_id: Optional[str] = Query(None),
    section: Optional[str] = Query(None),
    user=Depends(get_current_user)
):
    query = {}
    if exam_id:
        query["exam_id"] = exam_id
    
    if user["role"] == "student":
        session = await db.student_sessions.find_one({
            "student_id": user["id"],
            "status": "in_progress"
        })
        if session:
            sec_num = session.get("current_section", 1)
            query["section_number"] = {"$in": [sec_num, str(sec_num)]}
        else:
            raise HTTPException(status_code=403, detail="لا توجد جلسة نشطة مستمرة للحل")
    else:
        if section:
            try:
                sec_num = int(section)
                query["section_number"] = {"$in": [sec_num, str(sec_num)]}
            except ValueError:
                query["section_number"] = section

    questions = await db.questions.find(query, {"_id": 0}).to_list(1000)

    if user["role"] == "student":
        for q in questions:
            for opt in q.get("options", []):
                opt.pop("is_correct", None)
            if q.get("hint"):
                q["has_hint"] = True
                q.pop("hint", None)

    return questions

@api_router.get("/questions/{question_id}/hint")
async def get_hint(question_id: str, user=Depends(get_current_user)):
    if user["role"] != "student":
        raise HTTPException(status_code=403, detail="للطلاب فقط")

    question = await db.questions.find_one({"id": question_id}, {"_id": 0})
    if not question:
        raise HTTPException(status_code=404, detail="السؤال غير موجود")

    hint = question.get("hint")
    if not hint:
        return {"hint": None}

    return {"hint": hint}

@api_router.post("/questions")
async def create_question(data: QuestionCreate, user=Depends(get_current_user)):
    require_staff(user)
    question = {
        "id": str(uuid.uuid4()),
        "text": data.text,
        "exam_id": data.exam_id,
        "section_number": data.section_number,
        "points": data.points,
        "image": data.image,
        "hint": data.hint,
        "options": [o.model_dump() for o in data.options],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.questions.insert_one(question)
    question.pop("_id", None)
    return question

# ✅ ميزة حذف جميع الأسئلة بالكامل للباكند
@api_router.delete("/questions")
async def delete_all_questions(exam_id: Optional[str] = Query(None), user=Depends(get_current_user)):
    require_staff(user)
    query = {}
    if exam_id:
        query["exam_id"] = exam_id
    
    result = await db.questions.delete_many(query)
    return {"message": f"تم حذف جميع الأسئلة بنجاح (الإجمالي: {result.deleted_count})"}

@api_router.put("/questions/{question_id}")
async def update_question(question_id: str, data: QuestionUpdate, user=Depends(get_current_user)):
    require_staff(user)
    result = await db.questions.update_one({"id": question_id}, {"$set": {
        "text": data.text,
        "section_number": data.section_number,
        "points": data.points,
        "image": data.image,
        "hint": data.hint,
        "options": [o.model_dump() for o in data.options],
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="السؤال غير موجود")
    return {"message": "تم تحديث السؤال"}

@api_router.delete("/questions/{question_id}")
async def delete_question(question_id: str, user=Depends(get_current_user)):
    require_staff(user)
    result = await db.questions.delete_one({"id": question_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="السؤال غير موجود")
    return {"message": "تم حذف السؤال"}


# ================== STUDENT EXAM (كامل روابط الاختبار) ==================
@api_router.get("/student-exams")
async def get_student_exams(user=Depends(get_current_user)):
    if user["role"] == "student":
        query = {"student_id": user["id"]}
    else:
        query = {}
    sessions = await db.student_sessions.find(query, {"_id": 0}).to_list(1000)
    return sessions

@api_router.get("/student-exams/{session_id}")
async def get_student_exam(session_id: str, user=Depends(get_current_user)):
    session = await db.student_sessions.find_one({"id": session_id}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="جلسة الاختبار غير موجودة")
    return session

@api_router.post("/student-exams/start")
async def start_student_exam(data: StudentExamStart, user=Depends(get_current_user)):
    if user["role"] != "student":
        raise HTTPException(status_code=403, detail="للطلاب فقط")
    
    exam = await db.exams.find_one({"id": data.exam_id})
    if not exam:
        raise HTTPException(status_code=404, detail="الاختبار غير موجود")
    
    session = await db.student_sessions.find_one({
        "student_id": user["id"],
        "status": "in_progress"
    })
    if not session:
        raise HTTPException(status_code=404, detail="الجلسة غير موجودة")
    
    # تسجيل تاريخ بدء القسم الأول "1" فقط لحظة البدء الفعلية للاختبار ليعمل العداد التنازلي بدقة
    sections_start = {"1": datetime.now(timezone.utc).isoformat()}
    
    await db.student_sessions.update_one(
        {"id": session["id"]},
        {"$set": {
            "exam_id": data.exam_id,
            "section_start_times": sections_start,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    session["exam_id"] = data.exam_id
    session["section_start_times"] = sections_start
    session.pop("_id", None)
    return session

@api_router.post("/student-exams/answer")
async def submit_answer(data: StudentAnswerSubmit, user=Depends(get_current_user)):
    if user["role"] != "student":
        raise HTTPException(status_code=403, detail="للطلاب فقط")
    
    session = await db.student_sessions.find_one({"id": data.student_exam_id})
    if not session:
        raise HTTPException(status_code=404, detail="جلسة الاختبار غير موجودة")
    
    answers = session.get("answers", [])
    answers = [a for a in answers if a["question_id"] != data.question_id]
    
    answers.append({
        "question_id": data.question_id,
        "selected_option_id": data.selected_option_id,
        "section_number": session.get("current_section", 1)
    })
    
    await db.student_sessions.update_one(
        {"id": data.student_exam_id},
        {"$set": {"answers": answers}}
    )
    return {"message": "تم حفظ الإجابة بنجاح"}

@api_router.post("/student-exams/{session_id}/next-section")
async def next_section(session_id: str, user=Depends(get_current_user)):
    if user["role"] != "student":
        raise HTTPException(status_code=403, detail="للطلاب فقط")
    
    session = await db.student_sessions.find_one({"id": session_id})
    if not session:
        raise HTTPException(status_code=404, detail="جلسة الاختبار غير موجودة")
    
    current_section = session.get("current_section", 1)
    completed = session.get("completed_sections", [])
    if current_section not in completed:
        completed.append(current_section)
        
    next_sec = current_section + 1
    
    exam = await db.exams.find_one({"id": session["exam_id"]})
    total_sections = len(exam.get("sections", [])) if exam else 5
    
    is_completed = next_sec > total_sections
    
    # تحديث تواريخ بدء الأقسام ديناميكياً ولحظياً عند الدخول الفعلي للقسم الجديد لمسح تعليق العداد 00
    section_start_times = session.get("section_start_times", {})
    if not is_completed:
        section_start_times[str(next_sec)] = datetime.now(timezone.utc).isoformat()
    
    await db.student_sessions.update_one(
        {"id": session_id},
        {"$set": {
            "completed_sections": completed,
            "current_section": next_sec if not is_completed else current_section,
            "status": "completed" if is_completed else "in_progress",
            "section_start_times": section_start_times # حفظ تواريخ البدء اللحظية
        }}
    )
    return {"completed": is_completed, "next_section": next_sec}

@api_router.post("/student-exams/{session_id}/complete")
async def complete_student_exam(session_id: str, user=Depends(get_current_user)):
    if user["role"] != "student":
        raise HTTPException(status_code=403, detail="للطلاب فقط")
    
    result = await db.student_sessions.update_one(
        {"id": session_id},
        {"$set": {
            "status": "completed",
            "completed_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="جلسة الاختبار غير موجودة")
    return {"message": "تم إنهاء الاختبار بنجاح"}

@api_router.post("/student/submit-section")
async def submit_section(data: SectionSubmit, user=Depends(get_current_user)):
    if user["role"] != "student":
        raise HTTPException(status_code=403, detail="للطلاب فقط")

    session = await db.student_sessions.find_one({"id": data.session_id})
    if not session:
        raise HTTPException(status_code=404, detail="الجسلة غير موجودة")

    completed = session.get("completed_sections", [])

    if data.section_number in completed:
        raise HTTPException(status_code=400, detail="لا يمكن إعادة تقديم قسم مكتمل")

    questions = await db.questions.find({
        "exam_id": session.get("code"),
        "section_number": data.section_number
    }).to_list(1000)

    if len(questions) > 0:
        answered_ids = {a.question_id for a in data.answers}
        question_ids = {q["id"] for q in questions}
        if not question_ids.issubset(answered_ids):
            raise HTTPException(status_code=400, detail="يجب الإجابة على جميع الأسئلة قبل الانتقال")

    completed.append(data.section_number)
    answers = session.get("answers", [])
    for answer in data.answers:
        answers.append({
            "question_id": answer.question_id,
            "selected_option_id": answer.selected_option_id,
            "section_number": data.section_number
        })

    next_section = data.section_number + 1
    await db.student_sessions.update_one({"id": data.session_id}, {"$set": {
        "completed_sections": completed,
        "answers": answers,
        "current_section": next_section,
    }})

    return {"message": "تم تقديم القسم", "next_section": next_section}


# ================== GRADES ==================
@api_router.get("/grades")
async def get_grades(
    exam_id: Optional[str] = Query(None),
    user=Depends(get_current_user)
):
    require_staff(user)

    # 1. فلترة الجلسات بحسب معرف الاختبار
    query = {}
    if exam_id:
        query["exam_id"] = exam_id

    sessions = await db.student_sessions.find(query, {"_id": 0}).to_list(1000)

    # 2. حساب الإحصائيات العامة
    total = len(sessions)
    in_progress = sum(1 for s in sessions if s.get("status") == "in_progress")
    completed = sum(1 for s in sessions if s.get("status") == "completed")
    
    submissions = []
    graded_count = 0

    for session in sessions:
        score = 0
        total_points = 0
        answers = session.get("answers", [])

        # جلب جميع أسئلة هذا الاختبار لحساب المجموع الكلي الصحيح بدقة من الداتابيز 🚀
        exam_questions = await db.questions.find({"exam_id": session.get("exam_id")}).to_list(1000)
        total_points = sum(q.get("points", 1) for q in exam_questions)

        for answer in answers:
            # البحث عن السؤال المقابل في مصفوفة أسئلة الاختبار
            question = next((q for q in exam_questions if q["id"] == answer["question_id"]), None)
            if question:
                for opt in question.get("options", []):
                    if opt["id"] == answer["selected_option_id"] and opt["is_correct"]:
                        score += question.get("points", 1)

        # حساب النسبة المئوية الدقيقة (Skipped / Timed-out questions get 0 points) 🚀
        percentage = session.get("score")
        if percentage is None:
            percentage = round((score / total_points * 100) if total_points > 0 else 0, 1)
        
        is_graded = session.get("status") == "completed"
        if is_graded:
            graded_count += 1

        submissions.append({
            "id": session.get("id"),
            "student_name": session.get("student_name"),
            "grade": session.get("grade"),
            "class_name": session.get("class_name"),
            "school_name": session.get("school_name", ""), # قراءة مدرسة الطالب الحقيقية وإرسالها للواجهة (ستعرض فارغة إن تركها فارغة)
            "status": session.get("status"),
            "current_section": session.get("current_section", 1),
            "score": percentage if session.get("status") == "completed" else None,
            "started_at": session.get("started_at"),
        })

    return {
        "total": total,
        "in_progress": in_progress,
        "completed": completed,
        "graded": graded_count,
        "submissions": submissions
    }


@api_router.post("/grades/calculate/{session_id}")
async def calculate_student_grade(session_id: str, user=Depends(get_current_user)):
    require_staff(user)
    
    session = await db.student_sessions.find_one({"id": session_id})
    if not session:
        raise HTTPException(status_code=404, detail="جلسة الاختبار غير موجودة")
    
    # جلب جميع أسئلة هذا الاختبار لحساب المجموع الكلي الصحيح بدقة 🚀
    exam_questions = await db.questions.find({"exam_id": session.get("exam_id")}).to_list(1000)
    total_points = sum(q.get("points", 1) for q in exam_questions)
    
    score = 0
    answers = session.get("answers", [])
    for answer in answers:
        question = next((q for q in exam_questions if q["id"] == answer["question_id"]), None)
        if question:
            for opt in question.get("options", []):
                if opt["id"] == answer["selected_option_id"] and opt["is_correct"]:
                    score += question.get("points", 1)

    percentage = round((score / total_points * 100) if total_points > 0 else 0, 1)
    
    await db.student_sessions.update_one(
        {"id": session_id},
        {"$set": {"score": percentage, "total_points": total_points, "raw_score": score}}
    )
    
    return {"message": "تم حساب الدرجة بنجاح", "score": percentage}


@api_router.post("/grades/assign")
async def assign_student_grade(data: AssignGradeRequest, user=Depends(get_current_user)):
    require_staff(user)
    
    result = await db.student_sessions.update_one(
        {"id": data.student_exam_id},
        {"$set": {"score": data.score}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="جلسة الاختبار غير موجودة")
        
    return {"message": "تم تعيين الدرجة بنجاح"}


# ================== تنزيل الدرجات (تعديل الرصد اليدوي وإزالة الفصل) ==================
@api_router.get("/grades/export/excel")
async def export_grades_excel(
    exam_id: Optional[str] = Query(None),
    user=Depends(get_current_user)
):
    require_staff(user)

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="مكتبة openpyxl غير مثبتة. شغّل: pip install openpyxl")

    query = {"status": "completed"}
    if exam_id:
        query["exam_id"] = exam_id

    sessions = await db.student_sessions.find(query, {"_id": 0}).to_list(1000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "النتائج"
    ws.sheet_view.rightToLeft = True

    # حذف "الفصل" من ترويسة إكسل
    headers = ["اسم الطالب", "المدرسة", "المرحلة", "الدرجة", "المجموع", "النسبة %", "تاريخ الاختبار"]
    header_fill = PatternFill(start_color="3A7D86", end_color="3A7D86", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, session in enumerate(sessions, 2):
        exam_questions = await db.questions.find({"exam_id": session.get("exam_id")}).to_list(1000)
        total_points = sum(q.get("points", 1) for q in exam_questions)

        # التحقق مما إذا كانت الدرجة مرصودة يدوياً مسبقاً لكي لا تحسب القديمة
        percentage = session.get("score")
        if percentage is None:
            score = 0
            for answer in session.get("answers", []):
                question = next((q for q in exam_questions if q["id"] == answer["question_id"]), None)
                if question:
                    for opt in question.get("options", []):
                        if opt["id"] == answer["selected_option_id"] and opt["is_correct"]:
                            score += question.get("points", 1)
            percentage = round((score / total_points * 100) if total_points > 0 else 0, 1)
            raw_score_str = score
            total_points_str = total_points
        else:
            raw_score_str = session.get("raw_score", "-")
            total_points_str = session.get("total_points", "-")

        started = session.get("started_at", "")[:10]

        ws.cell(row=row_idx, column=1, value=session.get("student_name", ""))
        ws.cell(row=row_idx, column=2, value=session.get("school_name", "")) # تصدير مدرسة الطالب المكتوبة بدقة (وستعرض فارغة إن تركها فارغة)
        ws.cell(row=row_idx, column=3, value=session.get("grade", ""))
        ws.cell(row=row_idx, column=4, value=raw_score_str)
        ws.cell(row=row_idx, column=5, value=total_points_str) 
        ws.cell(row=row_idx, column=6, value=percentage)
        ws.cell(row=row_idx, column=7, value=started)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    output = io.BytesIO()
    wb.save(output)
    excel_data = output.getvalue()
    output.close()

    return Response(
        content=excel_data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=grades.xlsx"}
    )


@api_router.get("/grades/export/pdf")
async def export_grades_pdf(
    exam_id: Optional[str] = Query(None),
    user=Depends(get_current_user)
):
    require_staff(user)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    except ImportError:
        raise HTTPException(status_code=500, detail="مكتبة reportlab غير مثبتة. شغّل: pip install reportlab")

    query = {"status": "completed"}
    if exam_id:
        query["exam_id"] = exam_id

    sessions = await db.student_sessions.find(query, {"_id": 0}).to_list(1000)

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    elements = []

    # دالة ذكية لإعادة تشكيل النص العربي ليدعم الـ PDF بشكل صحيح
    def reshape_text(text: str) -> str:
        if not text:
            return ""
        try:
            import arabic_reshaper
            from bidi.algorithm import get_display
            reshaped = arabic_reshaper.reshape(text)
            return get_display(reshaped)
        except ImportError:
            return text

    # البحث الشامل عن خط عربي متوفر لكي لا يتلف الملف
    def register_arabic_font():
        try:
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import os
            
            font_paths = [
                "C:\\Windows\\Fonts\\tahoma.ttf",
                "C:\\Windows\\Fonts\\arial.ttf",
                "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
            ]
            for path in font_paths:
                if os.path.exists(path):
                    pdfmetrics.registerFont(TTFont('ArabicFont', path))
                    return 'ArabicFont'
        except Exception as e:
            print(f"Error registering font: {e}")
        return 'Helvetica'

    font_name = register_arabic_font()

    # تبديل الفصل بـ اسم المدرسة تلبية لطلبك وإظهارها في الـ PDF
    headers = [
        reshape_text("اسم الطالب"),
        reshape_text("المدرسة"),
        reshape_text("المرحلة"),
        reshape_text("الدرجة"),
        reshape_text("المجموع"),
        reshape_text("النسبة %")
    ]
    data = [headers]

    for session in sessions:
        # جلب جميع أسئلة هذا الاختبار لحساب المجموع الكلي الصحيح بدقة
        exam_questions = await db.questions.find({"exam_id": session.get("exam_id")}).to_list(1000)
        total_points = sum(q.get("points", 1) for q in exam_questions)

        # التحقق مما إذا كانت الدرجة مرصودة يدوياً مسبقاً لكي لا تحسب القديمة
        percentage = session.get("score")
        if percentage is None:
            score = 0
            for answer in session.get("answers", []):
                question = next((q for q in exam_questions if q["id"] == answer["question_id"]), None)
                if question:
                    for opt in question.get("options", []):
                        if opt["id"] == answer["selected_option_id"] and opt["is_correct"]:
                            score += question.get("points", 1)
            percentage = round((score / total_points * 100) if total_points > 0 else 0, 1)
            raw_score_str = str(score)
            total_points_str = str(total_points)
        else:
            raw_score_str = str(session.get("raw_score") if session.get("raw_score") is not None else "-")
            total_points_str = str(session.get("total_points") if session.get("total_points") is not None else "-")

        data.append([
            reshape_text(session.get("student_name", "")),
            reshape_text(session.get("school_name", "")), # تصدير مدرسة الطالب الحقيقية في الـ PDF (وستعرض فارغة إن تركها فارغة)
            reshape_text(session.get("grade", "")),
            raw_score_str,
            total_points_str,
            f"{percentage}%"
        ])

    # تعديل العروض colWidths لتتوافق مع الـ PDF المحدث وبخط ممركز
    table = Table(data, repeatRows=1, colWidths=[150, 130, 100, 40, 40, 60])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3A7D86")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    elements.append(table)
    doc.build(elements)
    pdf_data = output.getvalue()
    output.close()

    return Response(
        content=pdf_data,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=grades.pdf"}
    )


# ================== تصفير وحذف النتائج الكلي والمعلّق لضمان التصفير الكامل ==================
@api_router.delete("/results")
async def delete_results(data: DeleteResultsRequest, user=Depends(get_current_user)):
    require_staff(user)

    query = {}
    if data.exam_id:
        query["exam_id"] = data.exam_id
        await db.student_sessions.delete_many({"exam_id": {"$exists": False}})
        await db.student_sessions.delete_many({"exam_id": None})
    else:
        query = {}

    result = await db.student_sessions.delete_many(query)
    return {"message": f"تم تصفير وحذف جميع الجلسات والنتائج بنجاح (المحذوف: {result.deleted_count})"}


# ================== تصفير النظام بالكامل (تنظيف شامل للدفعة واستثناء n.vf11) ⚡ ==================
@api_router.post("/admin/super-reset")
async def super_reset(user=Depends(get_current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # 1. تصفير وحذف جميع الاختبارات، الأسئلة، الأكواد، الجلسات والطلاب بالكامل
    await db.exams.delete_many({})
    await db.questions.delete_many({})
    await db.student_codes.delete_many({})
    await db.student_sessions.delete_many({})
    
    # 2. حذف كافة المعلمين والمشرفين واستثناء المشرف الرئيسي n.vf11 بالملي 🔒
    await db.users.delete_many({"username": {"$ne": "n.vf11"}})
    
    return {"message": "✅ تم تصفير لوحة التحكم والنظام بالكامل بنجاح! تم حذف كافة الدفعات عدا حسابك الإشرافي الرئيسي."}


# ================== INIT ADMIN (فرض تهيئة حساب المشرف الرئيسي n.vf11) ==================
@api_router.post("/init-admin")
async def init_admin():
    # 1. حذف وتصفير أي حساب مشرف قديم أو تالف في قاعدة البيانات لضمان النظافة التامة
    await db.users.delete_many({"role": "admin"})
    await db.users.delete_many({"username": "admin"})
    await db.users.delete_many({"username": "n.vf11"})

    # 2. إنشاء حسابك الجديد المضمون والقوي 100% للأبد
    await db.users.insert_one({
        "id": str(uuid.uuid4()),
        "username": "n.vf11",
        "full_name": "المشرف الرئيسي",
        "role": "admin",
        "password": hash_password("Nn@100100"),
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    return {"message": "✅ تم إعادة تهيئة وإنشاء حساب المشرف الرئيسي الدائم بنجاح", "username": "n.vf11", "password": "Nn@100100"}

# ================== ROOT ==================
@api_router.get("/")
async def root():
    return {"message": "OK"}


app.include_router(api_router)